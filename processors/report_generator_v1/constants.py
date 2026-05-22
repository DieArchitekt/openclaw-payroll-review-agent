from openpyxl.styles import PatternFill

from gui.excel_styles import excel_fill, excel_font


CURRENCY_FIELDS: tuple[str, ...] = (
    "GrossPay",
    "PreTaxPension",
    "PAYE",
    "EmployeeNI",
    "PostTaxPension",
    "NetPay",
    "EmployerNI",
    "EmployerPension",
    "Current Value",
    "Previous Value",
)

CURRENCY_CONTAINS: tuple[str, ...] = (
    "GrossPay",
    "NetPay",
    "PAYE",
    "EmployerNI",
    "EmployerPension",
    "net pay",
    "employer cost",
)

SHEET_NAMES: tuple[str, ...] = (
    "Current Payroll",
    "Previous Payroll",
    "Reconciliation",
    "Anomalies",
    "Summary",
    "Current Field Recognition",
    "Previous Field Recognition",
)

HEADER_FILL = excel_fill("black")
HEADER_FONT = excel_font("text", bold=True)
HIGH_FILL = PatternFill("solid", fgColor="F4CCCC")
MEDIUM_FILL = PatternFill("solid", fgColor="FCE5CD")
SECTION_FILL = PatternFill("solid", fgColor="D9EAD3")
SUMMARY_LABEL_FILL = PatternFill("solid", fgColor="EFEFEF")
