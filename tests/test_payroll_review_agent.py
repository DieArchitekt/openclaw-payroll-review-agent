from io import BytesIO
import json
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook

from processors.anomaly_detector_v1 import detect_anomalies
import pytest

from processors.approval_workflow_v1 import (
    STATUS_APPROVED,
    STATUS_EXPORTED,
    STATUS_PREPARED,
    STATUS_QUERIES_RAISED,
    STATUS_REJECTED,
    STATUS_REVIEWED,
    approve_review,
    create_approval_record,
    mark_exported,
    mark_reviewed,
    raise_queries,
    reject_review,
)
from processors.openclaw_file_pairing import (
    discover_payroll_pairs,
    find_payroll_pair,
    wait_for_stable_payroll_pair,
)
from processors.openclaw_reporting import review_completion_message
from processors.payroll_processor_v1.extractor import extract_payroll
from processors.payroll_processor_v1.models import PayrollExtraction
from processors.payroll_review_cli import (
    cli_failure_message,
    output_prefix,
    run_cli,
    supported_file_types,
)
from processors.payroll_review_workflow import run_payroll_review
from processors.reconciliation_engine_v1 import reconcile_payroll
from processors.report_generator import generate_review_workbook


class UploadedPayrollFile:
    def __init__(self, name: str, payload: str) -> None:
        self.name = name
        self.payload = payload.encode("utf-8")

    def getvalue(self) -> bytes:
        return self.payload


def test_extract_payroll_from_csv_with_dynamic_headers(tmp_path):
    source = tmp_path / "current.csv"
    source.write_text(
        "Worker,Gross monthly,Tax deducted,National Insurance,Take home,Employer NI,Employers pension\n"
        "Ada Lovelace,3000,400,250,2350,300,150\n",
        encoding="utf-8",
    )

    extraction = extract_payroll(source)

    assert extraction.rows[0]["Employee"] == "Ada Lovelace"
    assert extraction.rows[0]["GrossPay"] == 3000.0
    assert extraction.rows[0]["PAYE"] == 400.0
    assert extraction.rows[0]["EmployeeNI"] == 250.0
    assert extraction.rows[0]["NetPay"] == 2350.0
    assert extraction.rows[0]["EmployerNI"] == 300.0
    assert extraction.rows[0]["EmployerPension"] == 150.0


def test_reconciliation_and_anomalies_flag_expected_rows():
    current_rows = [
        {
            "Employee": "Ada Lovelace",
            "GrossPay": 3600.0,
            "PAYE": 500.0,
            "NetPay": 2600.0,
            "EmployerNI": 320.0,
            "EmployerPension": 150.0,
        },
        {
            "Employee": "Grace Hopper",
            "GrossPay": 2500.0,
            "PAYE": 300.0,
            "NetPay": 2000.0,
            "EmployerNI": 250.0,
            "EmployerPension": 120.0,
        },
    ]
    previous_rows = [
        {
            "Employee": "Ada Lovelace",
            "GrossPay": 3000.0,
            "PAYE": 400.0,
            "NetPay": 2350.0,
            "EmployerNI": 300.0,
            "EmployerPension": 150.0,
        },
        {
            "Employee": "Alan Turing",
            "GrossPay": 2800.0,
            "PAYE": 350.0,
            "NetPay": 2200.0,
            "EmployerNI": 280.0,
            "EmployerPension": 130.0,
        },
    ]

    reconciliation_df, summary = reconcile_payroll(current_rows, previous_rows)
    anomalies_df = detect_anomalies(
        current_rows, reconciliation_df, summary, variance_threshold=10.0
    )

    assert set(reconciliation_df["Status"]) == {"Existing", "New", "Missing"}
    assert summary["new_employee_count"] == 1
    assert summary["missing_employee_count"] == 1
    assert "HIGH" in set(anomalies_df["Severity"])


def test_review_pack_contains_expected_sheets():
    current = PayrollExtraction(
        rows=[{"Employee": "Ada Lovelace", "GrossPay": 3000.0, "NetPay": 2350.0}]
    )
    previous = PayrollExtraction(
        rows=[{"Employee": "Ada Lovelace", "GrossPay": 2900.0, "NetPay": 2300.0}]
    )
    reconciliation_df, summary = reconcile_payroll(current.rows, previous.rows)
    anomalies_df = detect_anomalies(current.rows, reconciliation_df, summary)

    result = SimpleNamespace(
        current_extraction=current,
        previous_extraction=previous,
        reconciliation_df=reconciliation_df,
        anomalies_df=anomalies_df,
        summary=summary,
        approval_record=create_approval_record("preparer"),
    )

    payload = generate_review_workbook(result)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == [
        "Current Payroll",
        "Previous Payroll",
        "Reconciliation",
        "Anomalies",
        "Control Summary",
        "Summary",
        "Approval",
        "Current Field Recognition",
        "Previous Field Recognition",
    ]
    assert workbook["Reconciliation"].freeze_panes == "A2"
    assert workbook["Approval"]["A2"].value == "Review ID"
    assert workbook["Summary"]["A2"].value == "approval status"


def test_finance_grade_controls_detect_expected_exceptions():
    current_rows = [
        {
            "Employee": "Jane Smith",
            "BankAccount": "12345678",
            "NationalInsuranceNumber": "AB123456C",
            "GrossPay": 12000.0,
            "NetPay": 11000.0,
            "PAYE": 0.0,
            "EmployeeNI": 0.0,
            "EmployerPension": 0.0,
            "Department": "",
            "CostCentre": "",
            "BACSAmount": 11000.0,
        },
        {
            "Employee": "Jayne Smith",
            "BankAccount": "12345678",
            "NationalInsuranceNumber": "AB123456C",
            "GrossPay": 2000.0,
            "NetPay": 1500.0,
            "PAYE": 200.0,
            "EmployeeNI": 100.0,
            "EmployerPension": 80.0,
            "Department": "Operations",
            "CostCentre": "OPS",
            "BACSAmount": 1600.0,
        },
        {
            "Employee": "Leaver Person",
            "LeaverFlag": "Yes",
            "GrossPay": 1000.0,
            "NetPay": 800.0,
            "PAYE": 100.0,
            "EmployeeNI": 50.0,
            "EmployerPension": 40.0,
            "Department": "Finance",
            "CostCentre": "FIN",
            "BACSAmount": 800.0,
        },
        {
            "Employee": "Starter Person",
            "StarterFlag": "Yes",
            "StarterApproval": "",
            "GrossPay": 1000.0,
            "NetPay": 750.0,
            "PAYE": 100.0,
            "EmployeeNI": 50.0,
            "EmployerPension": 40.0,
            "Department": "Finance",
            "CostCentre": "FIN",
            "BACSAmount": 750.0,
        },
        {
            "Employee": "Negative Person",
            "GrossPay": 0.0,
            "NetPay": -50.0,
            "Department": "Finance",
            "CostCentre": "FIN",
            "BACSAmount": -50.0,
        },
    ]
    previous_rows = [
        {
            "Employee": "Jane Smith",
            "GrossPay": 1000.0,
            "Bonus": 100.0,
            "Overtime": 0.0,
            "Commission": 0.0,
        },
        {
            "Employee": "Jayne Smith",
            "GrossPay": 2000.0,
            "Bonus": 0.0,
            "Overtime": 0.0,
            "Commission": 0.0,
        },
    ]
    current_rows[0]["Bonus"] = 1000.0

    reconciliation_df, summary = reconcile_payroll(current_rows, previous_rows)
    anomalies_df = detect_anomalies(
        current_rows,
        reconciliation_df,
        summary,
        variance_threshold=20.0,
        high_net_pay_threshold=10000.0,
        bacs_tolerance=0.01,
    )
    categories = set(anomalies_df["Category"])

    assert "Duplicate Bank Account" in categories
    assert "Duplicate NI Number" in categories
    assert "Possible Duplicate Employee" in categories
    assert "High NetPay" in categories
    assert "Gross Pay With No Tax or NI" in categories
    assert "Employer Pension Missing" in categories
    assert "Missing Department" in categories
    assert "Missing CostCentre" in categories
    assert "Leaver Still Paid" in categories
    assert "Starter Approval Missing" in categories
    assert "Negative NetPay" in categories
    assert "BACS Control Difference" in categories
    assert "Variable Pay Movement" in categories


def test_run_payroll_review_returns_complete_result():
    current = UploadedPayrollFile(
        "current.csv",
        "Employee,GrossPay,PAYE,NetPay,EmployerNI,EmployerPension\nAda Lovelace,3000,400,2350,300,150\n",
    )
    previous = UploadedPayrollFile(
        "previous.csv",
        "Employee,GrossPay,PAYE,NetPay,EmployerNI,EmployerPension\nAda Lovelace,2900,390,2300,290,145\n",
    )

    result = run_payroll_review(
        current, previous, variance_threshold=20.0, prepared_by="Payroll preparer"
    )

    assert result.current_extraction.rows
    assert result.previous_extraction.rows
    assert not result.reconciliation_df.empty
    assert isinstance(result.summary, dict)
    assert result.approval_record.status == STATUS_PREPARED
    assert result.approval_record.prepared_by == "Payroll preparer"
    assert result.approval_record.review_id
    assert result.review_workbook_bytes.startswith(b"PK")


def test_approval_transitions_allow_expected_flow():
    record = create_approval_record("preparer")

    mark_reviewed(record, "reviewer", "looks fine")
    assert record.status == STATUS_REVIEWED
    assert record.reviewed_by == "reviewer"
    assert record.reviewer_comments == "looks fine"

    approve_review(record, "approver", "approved")
    assert record.status == STATUS_APPROVED
    assert record.approved_by == "approver"
    assert record.approval_comments == "approved"

    mark_exported(record, "exporter")
    assert record.status == STATUS_EXPORTED
    assert record.exported_by == "exporter"


def test_approval_queries_can_return_to_reviewed():
    record = create_approval_record("preparer")

    raise_queries(record, "reviewer", "missing explanation")
    assert record.status == STATUS_QUERIES_RAISED
    assert record.query_notes == "missing explanation"

    mark_reviewed(record, "reviewer", "query resolved")
    assert record.status == STATUS_REVIEWED
    assert record.reviewer_comments == "query resolved"


def test_approval_invalid_transitions_are_blocked():
    prepared = create_approval_record("preparer")

    with pytest.raises(ValueError):
        approve_review(prepared, "approver")

    rejected = create_approval_record("preparer")
    mark_reviewed(rejected, "reviewer")
    reject_review(rejected, "approver", "not acceptable")
    assert rejected.status == STATUS_REJECTED

    with pytest.raises(ValueError):
        mark_exported(rejected, "exporter")


def test_full_review_cli_writes_workbook_and_summary_json(tmp_path):
    current = tmp_path / "current.csv"
    previous = tmp_path / "previous.csv"
    output = tmp_path / "review.xlsx"
    summary_json = tmp_path / "summary.json"
    current.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,3000,400,250,2350,300,150\n",
        encoding="utf-8",
    )
    previous.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,2900,390,240,2300,290,145\n",
        encoding="utf-8",
    )

    exit_code = run_cli(
        [
            str(current),
            str(previous),
            "--out",
            str(output),
            "--summary-json",
            str(summary_json),
            "--prepared-by",
            "OpenClaw",
        ]
    )

    payload = json.loads(summary_json.read_text(encoding="utf-8"))
    assert exit_code == 0
    assert output.exists()
    assert output.read_bytes().startswith(b"PK")
    assert payload["approval_status"] == STATUS_PREPARED
    assert payload["prepared_by"] == "OpenClaw"
    assert payload["current_file"] == "current.csv"


def test_openclaw_folder_pairing_discovers_matching_current_previous_files(tmp_path):
    current_dir = tmp_path / "current"
    previous_dir = tmp_path / "previous"
    current_dir.mkdir()
    previous_dir.mkdir()
    current_file = current_dir / "client-a_2026-05_current.csv"
    previous_file = previous_dir / "client-a_2026-04_previous.csv"
    current_file.write_text("current", encoding="utf-8")
    previous_file.write_text("previous", encoding="utf-8")

    pair = find_payroll_pair(tmp_path)

    assert pair.key == "client-a"
    assert pair.current_path == current_file
    assert pair.previous_path == previous_file
    assert discover_payroll_pairs(tmp_path) == [pair]


def test_full_review_cli_can_use_incoming_root(tmp_path):
    incoming = tmp_path / "incoming_payroll"
    current_dir = incoming / "current"
    previous_dir = incoming / "previous"
    current_dir.mkdir(parents=True)
    previous_dir.mkdir(parents=True)
    output = tmp_path / "review.xlsx"
    summary_json = tmp_path / "summary.json"
    (current_dir / "payroll_current.csv").write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,3000,400,250,2350,300,150\n",
        encoding="utf-8",
    )
    (previous_dir / "payroll_previous.csv").write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,2900,390,240,2300,290,145\n",
        encoding="utf-8",
    )

    exit_code = run_cli(
        [
            "--incoming-root",
            str(incoming),
            "--out",
            str(output),
            "--summary-json",
            str(summary_json),
            "--prepared-by",
            "OpenClaw",
        ]
    )

    payload = json.loads(summary_json.read_text(encoding="utf-8"))
    assert exit_code == 0
    assert output.exists()
    assert payload["current_file"] == "payroll_current.csv"
    assert payload["previous_file"] == "payroll_previous.csv"


def test_wait_for_stable_payroll_pair_returns_existing_stable_pair(tmp_path):
    current_dir = tmp_path / "current"
    previous_dir = tmp_path / "previous"
    current_dir.mkdir()
    previous_dir.mkdir()
    current_file = current_dir / "payroll_current.csv"
    previous_file = previous_dir / "payroll_previous.csv"
    current_file.write_text("current", encoding="utf-8")
    previous_file.write_text("previous", encoding="utf-8")

    pair = wait_for_stable_payroll_pair(
        tmp_path,
        timeout_seconds=1,
        poll_interval_seconds=0.01,
        stable_checks=1,
    )

    assert pair.current_path == current_file
    assert pair.previous_path == previous_file


def test_full_review_cli_print_json_outputs_machine_readable_summary(tmp_path, capsys):
    current = tmp_path / "current.csv"
    previous = tmp_path / "previous.csv"
    output = tmp_path / "review.xlsx"
    summary_json = tmp_path / "summary.json"
    current.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,3000,400,250,2350,300,150\n",
        encoding="utf-8",
    )
    previous.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,2900,390,240,2300,290,145\n",
        encoding="utf-8",
    )

    exit_code = run_cli(
        [
            str(current),
            str(previous),
            "--out",
            str(output),
            "--summary-json",
            str(summary_json),
            "--print-json",
            "--prepared-by",
            "OpenClaw",
        ]
    )

    stdout_payload = json.loads(capsys.readouterr().out)
    file_payload = json.loads(summary_json.read_text(encoding="utf-8"))
    assert exit_code == 0
    assert stdout_payload["approval_status"] == STATUS_PREPARED
    assert stdout_payload["review_id"] == file_payload["review_id"]
    assert output.exists()


def test_cli_default_output_paths_are_named_and_do_not_overwrite(tmp_path):
    current = tmp_path / "client-a_2026-05_current.csv"
    previous = tmp_path / "client-a_2026-04_previous.csv"
    output_dir = tmp_path / "reviews"
    current.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,3000,400,250,2350,300,150\n",
        encoding="utf-8",
    )
    previous.write_text(
        "Employee,GrossPay,PAYE,EmployeeNI,NetPay,EmployerNI,EmployerPension\n"
        "Ada Lovelace,2900,390,240,2300,290,145\n",
        encoding="utf-8",
    )

    first_exit_code = run_cli(
        [
            str(current),
            str(previous),
            "--output-dir",
            str(output_dir),
            "--prepared-by",
            "OpenClaw",
        ]
    )
    second_exit_code = run_cli(
        [
            str(current),
            str(previous),
            "--output-dir",
            str(output_dir),
            "--prepared-by",
            "OpenClaw",
        ]
    )

    review_packs = sorted(output_dir.glob("client-a_2026-05*_review.xlsx"))
    summaries = sorted(output_dir.glob("client-a_2026-05*_summary.json"))
    assert first_exit_code == 0
    assert second_exit_code == 0
    assert len(review_packs) == 2
    assert len(summaries) == 2
    assert review_packs[0].read_bytes().startswith(b"PK")


def test_output_prefix_falls_back_to_timestamp_when_name_has_no_period():
    prefix = output_prefix(Path("current.csv"))

    assert prefix.startswith("payroll_review_")


def test_openclaw_completion_message_is_redacted():
    payload = {
        "review_id": "REV-1",
        "approval_status": "Prepared",
        "review_pack": "outputs/reviews/client-a_2026-05_review.xlsx",
        "high_exception_count": 2,
        "medium_exception_count": 3,
        "exception_count": 5,
        "summary": {"current_total_net_pay": 12345.67},
        "current_file": "client-a_2026-05_current.csv",
    }

    message = review_completion_message(payload)

    assert "Review ID: REV-1" in message
    assert "High exceptions: 2" in message
    assert "Medium exceptions: 3" in message
    assert "Total exceptions: 5" in message
    assert "Human review is required before approval/export." in message
    assert "12345.67" not in message
    assert "client-a_2026-05_current.csv" not in message


def test_cli_rejects_unsupported_explicit_input_file(tmp_path):
    current = tmp_path / "current.docx"
    previous = tmp_path / "previous.csv"
    current.write_text("not supported", encoding="utf-8")
    previous.write_text("Employee,NetPay\nAda Lovelace,2350\n", encoding="utf-8")

    with pytest.raises(SystemExit) as exc:
        run_cli([str(current), str(previous)])

    assert "unsupported file type" in str(exc.value)
    assert supported_file_types() in str(exc.value)


def test_cli_failure_message_is_short_and_non_destructive():
    message = cli_failure_message("Current payroll file not found: missing.csv")

    assert "Payroll review failed." in message
    assert "Current payroll file not found" in message
    assert "No files were moved" in message
