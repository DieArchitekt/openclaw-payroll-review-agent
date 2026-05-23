from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from gui.excel_styles import fill_row_cells, money_format

from .constants import (
    CURRENCY_CONTAINS,
    CURRENCY_FIELDS,
    HEADER_FILL,
    HEADER_FONT,
    HIGH_FILL,
    MEDIUM_FILL,
    SECTION_FILL,
    SUMMARY_LABEL_FILL,
)


def style_header_row(ws: Worksheet) -> None:
    """Apply bold dark formatting to the top row."""
    if ws.max_row < 1:
        return

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def style_anomaly_sheet(ws: Worksheet) -> None:
    """Highlight HIGH and MEDIUM anomaly rows."""
    severity_column: int | None = find_column(ws, "Severity")

    if not severity_column:
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        severity = str(row[severity_column - 1].value or "").upper()

        if severity == "HIGH":
            fill_row(row, HIGH_FILL)

        if severity == "MEDIUM":
            fill_row(row, MEDIUM_FILL)


def auto_size_columns(ws: Worksheet) -> None:
    """Auto-size worksheet columns based on visible content."""
    for column_cells in ws.columns:
        values = [
            "" if cell.value is None else str(cell.value) for cell in column_cells
        ]
        width = min(max(max((len(value) for value in values), default=0) + 2, 12), 42)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def write_summary_metrics(ws: Worksheet) -> None:
    """Apply summary-specific section formatting."""
    ws.freeze_panes = None
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        metric = str(row[0].value or "")
        value = row[1].value

        if value in ("", None):
            fill_row(row, SECTION_FILL)
            row[0].font = Font(bold=True)
        else:
            row[0].fill = SUMMARY_LABEL_FILL
            row[0].font = Font(bold=True)

        if is_money_metric(metric):
            money_format(row[1])

        if metric.endswith("%"):
            row[1].number_format = "0.00"


def apply_currency_formats(ws: Worksheet) -> None:
    """Apply two-decimal formatting to currency-like columns."""
    for col_idx, header in header_map(ws).items():
        if is_currency_header(header):
            for row in ws.iter_rows(
                min_row=2, min_col=col_idx, max_col=col_idx, max_row=ws.max_row
            ):
                money_format(row[0])


def format_standard_sheet(ws: Worksheet) -> None:
    """Apply shared formatting to a worksheet."""
    style_header_row(ws)
    ws.freeze_panes = "A2"
    apply_currency_formats(ws)
    auto_size_columns(ws)


def header_map(ws: Worksheet) -> dict[int, str]:
    """Return worksheet header names by one-based column index."""
    if ws.max_row < 1:
        return {}

    return {cell.column: str(cell.value or "") for cell in ws[1]}


def find_column(ws: Worksheet, header: str) -> int | None:
    """Return the one-based column index for a header."""
    for col_idx, value in header_map(ws).items():
        if value == header:
            return col_idx

    return None


def is_currency_header(header: str) -> bool:
    """Return whether a column header should receive currency formatting."""
    return header in CURRENCY_FIELDS or any(
        token in header for token in CURRENCY_CONTAINS
    )


def is_money_metric(metric: str) -> bool:
    """Return whether a summary metric is money-like."""
    metric = metric.lower()
    return any(
        token in metric for token in ("net pay", "employer cost")
    ) and not metric.endswith("%")


def fill_row(row, fill: PatternFill) -> None:
    """Apply a fill to every cell in a row."""
    fill_row_cells(row, fill)
