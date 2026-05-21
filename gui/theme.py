THEME = {
    "black": "#050507",
    "panel": "#101014",
    "field": "#171721",
    "border": "#2a1f3d",
    "text": "#f5f1ff",
    "muted": "#a79abf",
    "magenta": "#ff2bd6",
    "ultraviolet": "#7b2cff",
    "disabled": "#332944",
}

EXCEL_THEME = {
    "black": "050507",
    "text": "F5F1FF",
    "magenta": "FF2BD6",
    "ultraviolet": "7B2CFF",
}

RESET = "\033[0m"
TERMINAL_MAGENTA = "\033[95m"
TERMINAL_ULTRAVIOLET = "\033[38;5;99m"


def palette():
    return dict(THEME)


def color(name):
    return THEME[name]


def important(text):
    return f"{TERMINAL_MAGENTA}{text}{RESET}"


def accent(text):
    return f"{TERMINAL_ULTRAVIOLET}{text}{RESET}"


def excel_fills():
    from openpyxl.styles import PatternFill

    return {
        "black": PatternFill("solid", fgColor=EXCEL_THEME["black"]),
        "magenta": PatternFill("solid", fgColor=EXCEL_THEME["magenta"]),
        "ultraviolet": PatternFill("solid", fgColor=EXCEL_THEME["ultraviolet"]),
    }


def excel_color(name):
    return EXCEL_THEME[name]


def label_style(background="panel", foreground="text"):
    return {
        "bg": THEME[background],
        "fg": THEME[foreground],
    }


def entry_style():
    return {
        "bg": THEME["field"],
        "fg": THEME["text"],
        "insertbackground": THEME["magenta"],
        "relief": "flat",
    }


def button_style(background="ultraviolet", foreground="text"):
    active_background = "magenta"

    if background == "magenta":
        active_background = "ultraviolet"

    return {
        "bg": THEME[background],
        "fg": THEME[foreground],
        "activebackground": THEME[active_background],
        "activeforeground": THEME["black"],
        "relief": "flat",
    }


def streamlit_css():
    return f"""
<style>
:root {{
    --payroll-black: {THEME["black"]};
    --payroll-panel: {THEME["panel"]};
    --payroll-field: {THEME["field"]};
    --payroll-border: {THEME["border"]};
    --payroll-text: {THEME["text"]};
    --payroll-muted: {THEME["muted"]};
    --payroll-magenta: {THEME["magenta"]};
    --payroll-ultraviolet: {THEME["ultraviolet"]};
}}

.stApp {{
    background: var(--payroll-black);
    color: var(--payroll-text);
}}

[data-testid="stSidebar"],
[data-testid="stHeader"] {{
    background: var(--payroll-panel);
}}

.stButton > button,
.stDownloadButton > button {{
    background: var(--payroll-magenta);
    border: 1px solid var(--payroll-magenta);
    color: var(--payroll-black);
    font-weight: 700;
}}

.stButton > button:hover,
.stDownloadButton > button:hover {{
    background: var(--payroll-ultraviolet);
    border-color: var(--payroll-ultraviolet);
    color: var(--payroll-text);
}}

input,
textarea,
[data-baseweb="select"] {{
    background: var(--payroll-field);
    color: var(--payroll-text);
}}
</style>
"""


def use_streamlit_theme(st):
    st.markdown(streamlit_css(), unsafe_allow_html=True)
