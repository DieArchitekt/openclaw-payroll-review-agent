from openpyxl.cell.cell import Cell
from openpyxl.styles import Font, PatternFill, numbers

from .theme import excel_color


def excel_fill(name: str) -> PatternFill:
    return PatternFill("solid", fgColor=excel_color(name))


def excel_font(name: str = "text", bold: bool = False) -> Font:
    return Font(bold=bold, color=excel_color(name))


def money_format(cell: Cell) -> None:
    cell.number_format = numbers.FORMAT_NUMBER_00


def fill_row_cells(row, fill: PatternFill) -> None:
    for cell in row:
        cell.fill = fill
