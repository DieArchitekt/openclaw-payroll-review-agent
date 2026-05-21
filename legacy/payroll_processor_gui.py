import re
import sys
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

ROOT_DIR = Path(__file__).resolve().parents[1]

if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from gui.theme import button_style, color, entry_style, excel_color, excel_fills, label_style


HEADERS = [
    "Employee",
    "GrossMonthly",
    "PreTaxPen",
    "PAYE",
    "EesNI",
    "PostTaxPen",
    "NetPay",
    "ErsNI",
    "ErsPen",
]

def parse_value(value):
    if not value or value.strip() == "":
        return 0.0

    try:
        return float(value.replace(",", ""))
    except (AttributeError, ValueError):
        return 0.0


def extract_payroll_from_pdf(pdf_path):
    employees = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                if should_skip_line(line):
                    continue

                match = re.match(r"^(\d+)\s+(.+?)\s+([A-Z])\s+([\d.\s]+)$", line.strip())
                if not match:
                    continue

                _, name_and_tax_code, _, numbers_text = match.groups()
                name = clean_employee_name(name_and_tax_code)
                values = re.findall(r"\d+\.\d+", numbers_text)

                if not is_valid_employee_name(name) or len(values) < 15:
                    continue

                employees.append(
                    {
                        "Employee": name,
                        "GrossMonthly": parse_value(values[0]),
                        "PreTaxPen": parse_value(values[4]),
                        "PAYE": parse_value(values[6]),
                        "EesNI": parse_value(values[7]),
                        "PostTaxPen": parse_value(values[9]),
                        "NetPay": parse_value(values[12]),
                        "ErsNI": parse_value(values[13]),
                        "ErsPen": parse_value(values[14]),
                    }
                )

    return employees


def should_skip_line(line):
    if not line:
        return True

    skipped_terms = (
        "Employee",
        "Totals",
        "Frequency",
        "Pay period",
        "Run date",
        "Utheo",
        "Limited",
        "Tax Code",
    )

    return "TOTAL" in line.upper() or any(term in line for term in skipped_terms)


def clean_employee_name(name):
    return re.sub(r"S[A-Z0-9]+[RLM]+$", "", name).strip()


def is_valid_employee_name(name):
    return bool(name and len(name) >= 3 and any(char.isalpha() for char in name))


def create_excel_with_spacing(employees, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    fills = excel_fills()

    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=excel_color("text"))
        cell.fill = fills["black"]
        cell.alignment = Alignment(horizontal="center")

    data_rows = []
    current_row = 2

    for employee in employees:
        for col_idx, header in enumerate(HEADERS, start=1):
            value = employee[header]
            cell = ws.cell(row=current_row, column=col_idx, value=value)

            if col_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER_00

            if col_idx == 1:
                cell.font = Font(color=excel_color("magenta"))

        data_rows.append(current_row)
        current_row += 2

    totals_row = current_row
    totals_label = ws.cell(row=totals_row, column=1, value="Totals")
    totals_label.font = Font(bold=True, color=excel_color("black"))
    totals_label.fill = fills["magenta"]

    first_data_row = data_rows[0]
    last_data_row = data_rows[-1]

    for col_idx in range(2, len(HEADERS) + 1):
        col_letter = chr(64 + col_idx)
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell.font = Font(bold=True, color=excel_color("text"))
        cell.fill = fills["ultraviolet"]
        cell.number_format = numbers.FORMAT_NUMBER_00

    ws.column_dimensions["A"].width = 25

    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)


class PayrollProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Payroll Processor")
        self.root.geometry("700x450")
        self.root.resizable(False, False)
        self.root.configure(bg=color("black"))

        self.pdf_path = tk.StringVar()
        self.output_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self):
        title_frame = tk.Frame(self.root, bg=color("black"), height=80)
        title_frame.pack(fill=tk.X, pady=0)
        title_frame.pack_propagate(False)

        title_label = tk.Label(
            title_frame,
            text="Payroll PDF to Excel Converter",
            font=("Arial", 20, "bold"),
            **label_style("black", "magenta"),
        )
        title_label.pack(pady=20)

        subtitle_label = tk.Label(
            title_frame,
            text="Convert payroll invoices to formatted Excel spreadsheets",
            font=("Arial", 10),
            **label_style("black", "muted"),
        )
        subtitle_label.pack(pady=0)

        main_frame = tk.Frame(self.root, padx=30, pady=20, bg=color("panel"))
        main_frame.pack(fill=tk.BOTH, expand=True)

        input_label = tk.Label(
            main_frame,
            text="1. Select your PDF invoice file:",
            font=("Arial", 11, "bold"),
            **label_style(),
        )
        input_label.grid(row=0, column=0, sticky="w", pady=(10, 5))

        input_frame = tk.Frame(main_frame, bg=color("panel"))
        input_frame.grid(row=1, column=0, sticky="ew", pady=(0, 20))

        input_entry = tk.Entry(
            input_frame,
            textvariable=self.pdf_path,
            font=("Arial", 10),
            width=50,
            state="readonly",
            readonlybackground=color("field"),
            **entry_style(),
        )
        input_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        input_browse_btn = tk.Button(
            input_frame,
            text="Browse PDF...",
            command=self.browse_pdf,
            font=("Arial", 10, "bold"),
            cursor="hand2",
            width=15,
            **button_style(),
        )
        input_browse_btn.pack(side=tk.RIGHT)

        output_label = tk.Label(
            main_frame,
            text="2. Choose where to save the Excel file:",
            font=("Arial", 11, "bold"),
            **label_style(),
        )
        output_label.grid(row=2, column=0, sticky="w", pady=(10, 5))

        output_frame = tk.Frame(main_frame, bg=color("panel"))
        output_frame.grid(row=3, column=0, sticky="ew", pady=(0, 20))

        output_entry = tk.Entry(
            output_frame,
            textvariable=self.output_path,
            font=("Arial", 10),
            width=50,
            **entry_style(),
        )
        output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        output_browse_btn = tk.Button(
            output_frame,
            text="Browse Folder...",
            command=self.browse_output,
            font=("Arial", 10, "bold"),
            cursor="hand2",
            width=15,
            **button_style(),
        )
        output_browse_btn.pack(side=tk.RIGHT)

        hint_label = tk.Label(
            main_frame,
            text="Tip: You can type a custom filename like 'feb_invoices' in the box above",
            font=("Arial", 9),
            **label_style("panel", "muted"),
        )
        hint_label.grid(row=4, column=0, sticky="w", pady=(0, 20))

        self.process_btn = tk.Button(
            main_frame,
            text="Convert PDF to Excel",
            command=self.process_payroll,
            font=("Arial", 14, "bold"),
            cursor="hand2",
            height=2,
            **button_style("magenta", "black"),
        )
        self.process_btn.grid(row=5, column=0, sticky="ew", pady=(20, 10))

        self.status_label = tk.Label(
            main_frame,
            text="Ready to convert your payroll PDF.",
            font=("Arial", 10),
            **label_style("panel", "muted"),
            wraplength=600,
            justify="left",
        )
        self.status_label.grid(row=6, column=0, sticky="w", pady=(10, 0))

        main_frame.columnconfigure(0, weight=1)

    def browse_pdf(self):
        filename = filedialog.askopenfilename(
            title="Select PDF Invoice File",
            filetypes=[
                ("PDF files", "*.pdf"),
                ("All files", "*.*"),
            ],
        )

        if not filename:
            return

        self.pdf_path.set(filename)

        pdf_path = Path(filename)
        suggested_output = pdf_path.parent / f"{pdf_path.stem}_processed.xlsx"
        self.output_path.set(str(suggested_output))
        self.status_label.config(text=f"Selected: {pdf_path.name}", fg=color("magenta"))

    def browse_output(self):
        initial_file = "payroll.xlsx"

        if self.output_path.get():
            initial_file = Path(self.output_path.get()).name

        filename = filedialog.asksaveasfilename(
            title="Save Excel File As",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel files", "*.xlsx"),
                ("All files", "*.*"),
            ],
            initialfile=initial_file,
        )

        if not filename:
            return

        self.output_path.set(filename)
        self.status_label.config(text=f"Will save to: {Path(filename).name}", fg=color("magenta"))

    def process_payroll(self):
        pdf_file = self.pdf_path.get()
        output_file = self.output_path.get()

        if not pdf_file:
            messagebox.showerror(
                "Missing Input",
                "Please select a PDF file first.\n\nClick 'Browse PDF...' to choose your invoice file.",
            )
            return

        if not output_file:
            messagebox.showerror(
                "Missing Output",
                "Please specify where to save the Excel file.\n\nClick 'Browse Folder...' or type a filename.",
            )
            return

        if not Path(pdf_file).exists():
            messagebox.showerror(
                "File Not Found",
                f"The PDF file does not exist:\n\n{pdf_file}\n\nPlease select a valid file.",
            )
            return

        output_path = Path(output_file)

        if output_path.suffix.lower() != ".xlsx":
            output_path = output_path.with_suffix(".xlsx")
            self.output_path.set(str(output_path))

        try:
            self.process_btn.config(state="disabled", bg=color("disabled"))
            self.status_label.config(text="Processing PDF, please wait...", fg=color("ultraviolet"))
            self.root.update()

            employees = extract_payroll_from_pdf(Path(pdf_file))

            if not employees:
                messagebox.showerror(
                    "No Data Found",
                    "Could not find any employee data in the PDF.\n\n"
                    "Please make sure this is a valid payroll invoice PDF.",
                )
                self.status_label.config(
                    text="Failed: No employee data found in PDF",
                    fg=color("magenta"),
                )
                return

            create_excel_with_spacing(employees, output_path)

            messagebox.showinfo(
                "Success",
                f"Payroll processed successfully.\n\n"
                f"Found {len(employees)} employees\n"
                f"Created Excel with totals\n"
                f"Saved to: {output_path.name}",
            )

            self.status_label.config(
                text=f"Success. Processed {len(employees)} employees to {output_path.name}",
                fg=color("magenta"),
            )

        except Exception as exc:
            messagebox.showerror(
                "Error",
                f"An error occurred while processing:\n\n{exc}\n\n"
                "Please check the PDF file and try again.",
            )
            self.status_label.config(text=f"Error: {exc}", fg=color("magenta"))

        finally:
            self.process_btn.config(state="normal", bg=color("magenta"))


def main():
    root = tk.Tk()
    PayrollProcessorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
