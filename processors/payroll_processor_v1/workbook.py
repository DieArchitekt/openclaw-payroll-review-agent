from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from gui.excel_styles import excel_fill, excel_font, money_format
from .models import FieldMatch, PayrollExtraction
from .schema import EXPORT_FIELDS, PAYROLL_SCHEMA, output_header


def exported_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Return rows reduced to configured output fields only."""
    return [{output_header(field): output_value(row, field) for field in EXPORT_FIELDS} for row in rows]


def output_value(row: dict[str, Any], field_name: str) -> Any:
    """Return an exportable value for a canonical field."""
    default: Any = 0.0 if PAYROLL_SCHEMA[field_name]["kind"] in {"money", "number"} else ""
    return row.get(field_name, default)


def create_payroll_workbook(extraction: PayrollExtraction) -> Workbook:
    """Return an XLSX workbook for the extracted payroll data."""
    if not extraction.rows:
        raise ValueError("No payroll rows were found in the payroll file.")

    wb: Workbook = Workbook()
    add_export_sheet(wb, extraction.rows)
    add_recognition_sheet(wb, extraction)

    return wb


def save_payroll_workbook(extraction: PayrollExtraction, output_path: Path) -> None:
    """Save extracted payroll data to an XLSX file."""
    create_payroll_workbook(extraction).save(output_path)


def workbook_to_bytes(extraction: PayrollExtraction) -> bytes:
    """Return extracted payroll data as XLSX bytes for downloads."""
    buffer: BytesIO = BytesIO()
    create_payroll_workbook(extraction).save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def add_export_sheet(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    """Create the main payroll export sheet."""
    ws: Worksheet = wb.active
    ws.title = "Payroll"
    headers: list[str] = [output_header(field_name) for field_name in EXPORT_FIELDS]

    write_header_row(ws, headers)
    data_rows: list[int] = write_export_rows(ws, headers, rows)
    write_totals_row(ws, headers, data_rows)
    size_columns(ws, headers)


def write_header_row(ws: Worksheet, headers: list[str]) -> None:
    """Write the styled header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell: Cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell)


def write_export_rows(ws: Worksheet, headers: list[str], rows: list[dict[str, Any]]) -> list[int]:
    """Write exported payroll rows and return their row numbers."""
    data_rows: list[int] = []
    current_row: int = 2

    for row in exported_rows(rows):
        write_export_row(ws, current_row, headers, row)
        data_rows.append(current_row)
        current_row += 2

    return data_rows


def write_export_row(ws: Worksheet, row_number: int, headers: list[str], row: dict[str, Any]) -> None:
    """Write one exported payroll row."""
    for col_idx, header in enumerate(headers, start=1):
        value: Any = row.get(header, "")
        cell: Cell = ws.cell(row=row_number, column=col_idx, value=value)
        style_data_cell(cell, col_idx, value)


def write_totals_row(ws: Worksheet, headers: list[str], data_rows: list[int]) -> None:
    """Write totals formulas for exported numeric columns."""
    if not data_rows:
        return

    totals_row: int = data_rows[-1] + 2
    first_data_row: int = data_rows[0]
    last_data_row: int = data_rows[-1]
    totals_label: Cell = ws.cell(row=totals_row, column=1, value="Totals")
    style_total_label(totals_label)

    for col_idx in range(2, len(headers) + 1):
        write_total_formula(ws, totals_row, col_idx, first_data_row, last_data_row)


def write_total_formula(ws: Worksheet, totals_row: int, col_idx: int, first_row: int, last_row: int) -> None:
    """Write one Excel SUM formula."""
    col_letter: str = chr(64 + col_idx)
    cell: Cell = ws.cell(row=totals_row, column=col_idx)
    cell.value = f"=SUM({col_letter}{first_row}:{col_letter}{last_row})"
    style_total_value(cell)


def add_recognition_sheet(wb: Workbook, extraction: PayrollExtraction) -> None:
    """Create an audit sheet showing recognised, ignored, and unmapped fields."""
    ws: Worksheet = wb.create_sheet("Field Recognition")
    headers: list[str] = ["Source Header", "Canonical Field", "Status", "Confidence", "Reason"]
    write_header_row(ws, headers)

    for row_idx, match in enumerate(extraction.field_matches, start=2):
        write_recognition_row(ws, row_idx, match)

    for col_letter in "ABCDE":
        ws.column_dimensions[col_letter].width = 24


def write_recognition_row(ws: Worksheet, row_idx: int, match: FieldMatch) -> None:
    """Write one field-recognition audit row."""
    values: list[Any] = [match.source_header, match.canonical_field or "", match.status, match.confidence, match.reason]

    for col_idx, value in enumerate(values, start=1):
        cell: Cell = ws.cell(row=row_idx, column=col_idx, value=value)
        style_data_cell(cell, col_idx, value)


def size_columns(ws: Worksheet, headers: list[str]) -> None:
    """Set readable widths for export columns."""
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = max(14, min(28, len(header) + 4))


def style_header_cell(cell: Cell) -> None:
    """Apply workbook styling to one header cell."""
    cell.font = excel_font("text", bold=True)
    cell.fill = excel_fill("black")
    cell.alignment = Alignment(horizontal="center")


def style_data_cell(cell: Cell, column_index: int, value: Any) -> None:
    """Apply workbook styling to one data cell."""
    if column_index > 1 and isinstance(value, (int, float)):
        money_format(cell)

    if column_index == 1:
        cell.font = excel_font("magenta")


def style_total_label(cell: Cell) -> None:
    """Apply workbook styling to the totals label."""
    cell.font = excel_font("black", bold=True)
    cell.fill = excel_fill("magenta")


def style_total_value(cell: Cell) -> None:
    """Apply workbook styling to one totals formula cell."""
    cell.font = excel_font("text", bold=True)
    cell.fill = excel_fill("ultraviolet")
    money_format(cell)
