from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from processors.anomaly_detector_v1 import detect_anomalies
from processors.approval_workflow_v1 import create_approval_record
from processors.payroll_processor_v1.models import PayrollExtraction
from processors.reconciliation_engine_v1 import reconcile_payroll
from processors.report_generator_v1 import generate_review_workbook


def test_review_pack_contains_expected_sheets():
    current = PayrollExtraction(
        rows=[{"Employee": "Ada Lovelace", "GrossPay": 3000.0, "NetPay": 2350.0}]
    )
    previous = PayrollExtraction(
        rows=[{"Employee": "Ada Lovelace", "GrossPay": 2900.0, "NetPay": 2300.0}]
    )
    reconciliation_df, summary = reconcile_payroll(current.rows, previous.rows)
    anomalies_df = detect_anomalies(current.rows, reconciliation_df, summary)

    result = SimpleNamespace(
        current_extraction=current,
        previous_extraction=previous,
        reconciliation_df=reconciliation_df,
        anomalies_df=anomalies_df,
        summary=summary,
        approval_record=create_approval_record("preparer"),
    )

    payload = generate_review_workbook(result)
    workbook = load_workbook(BytesIO(payload))

    assert workbook.sheetnames == [
        "Current Payroll",
        "Previous Payroll",
        "Reconciliation",
        "Anomalies",
        "Control Summary",
        "Summary",
        "Approval",
        "Run Metadata",
        "Current Field Recognition",
        "Previous Field Recognition",
    ]
    assert workbook["Reconciliation"].freeze_panes == "A2"
    assert workbook["Approval"]["A2"].value == "Review ID"
    assert workbook["Summary"]["A2"].value == "approval status"
    assert workbook["Run Metadata"]["A2"].value == "Schema version"
