import argparse
import re
import sys
from pathlib import Path

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, numbers

ROOT_DIR = Path(__file__).resolve().parents[1]

if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from gui.theme import accent, excel_color, excel_fills, important


PDF_TO_EXCEL_MAP = {
    1: ("A", "Employee"),
    3: ("B", "GrossMonthly"),
    7: ("C", "PreTaxPen"),
    9: ("D", "PAYE"),
    10: ("E", "EesNI"),
    12: ("F", "PostTaxPen"),
    15: ("G", "NetPay"),
    16: ("H", "ErsNI"),
    17: ("I", "ErsPen"),
}

def parse_value(value):
    if not value or value.strip() == "":
        return 0.0

    try:
        return float(value.replace(",", ""))
    except (AttributeError, ValueError):
        return 0.0


def extract_payroll_from_pdf(pdf_path):
    employees = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue

            for line in text.split("\n"):
                if should_skip_line(line):
                    continue

                match = re.match(r"^(\d+)\s+(.+?)\s+([A-Z])\s+([\d.\s]+)$", line.strip())
                if not match:
                    continue

                _, name_and_tax_code, _, numbers_text = match.groups()
                name = clean_employee_name(name_and_tax_code)
                values = re.findall(r"\d+\.\d+", numbers_text)

                if not is_valid_employee_name(name) or len(values) < 15:
                    continue

                employees.append(
                    {
                        "Employee": name,
                        "GrossMonthly": parse_value(values[0]),
                        "PreTaxPen": parse_value(values[4]),
                        "PAYE": parse_value(values[6]),
                        "EesNI": parse_value(values[7]),
                        "PostTaxPen": parse_value(values[9]),
                        "NetPay": parse_value(values[12]),
                        "ErsNI": parse_value(values[13]),
                        "ErsPen": parse_value(values[14]),
                    }
                )

    return employees


def should_skip_line(line):
    if not line:
        return True

    skipped_terms = (
        "Employee",
        "Totals",
        "Frequency",
        "Pay period",
        "Run date",
        "Utheo",
        "Limited",
        "Tax Code",
    )

    return "TOTAL" in line.upper() or any(term in line for term in skipped_terms)


def clean_employee_name(name):
    return re.sub(r"S[A-Z0-9]+[RLM]+$", "", name).strip()


def is_valid_employee_name(name):
    return bool(name and len(name) >= 3 and any(char.isalpha() for char in name))


def create_excel_with_spacing(employees, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    fills = excel_fills()

    headers = [header for _, header in sorted(PDF_TO_EXCEL_MAP.values(), key=lambda item: item[0])]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=excel_color("text"))
        cell.fill = fills["black"]
        cell.alignment = Alignment(horizontal="center")

    data_rows = []
    current_row = 2

    for employee in employees:
        for col_idx, header in enumerate(headers, start=1):
            value = employee[header]
            cell = ws.cell(row=current_row, column=col_idx, value=value)

            if col_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER_00

            if col_idx == 1:
                cell.font = Font(color=excel_color("magenta"))

        data_rows.append(current_row)
        current_row += 2

    totals_row = current_row
    totals_label = ws.cell(row=totals_row, column=1, value="Totals")
    totals_label.font = Font(bold=True, color=excel_color("black"))
    totals_label.fill = fills["magenta"]

    first_data_row = data_rows[0]
    last_data_row = data_rows[-1]

    for col_idx in range(2, len(headers) + 1):
        col_letter = chr(64 + col_idx)
        cell = ws.cell(row=totals_row, column=col_idx)
        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
        cell.font = Font(bold=True, color=excel_color("text"))
        cell.fill = fills["ultraviolet"]
        cell.number_format = numbers.FORMAT_NUMBER_00

    ws.column_dimensions["A"].width = 25

    for col_letter in "BCDEFGHI":
        ws.column_dimensions[col_letter].width = 14

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Process a payroll PDF invoice into a formatted Excel file."
    )
    parser.add_argument("pdf", type=Path, help="Input PDF invoice file")
    parser.add_argument("-o", "--out", type=Path, help="Output XLSX file")

    args = parser.parse_args()

    if not args.pdf.exists():
        raise SystemExit(f"PDF file not found: {args.pdf}")

    print(f"{accent('Reading PDF:')} {args.pdf}")
    employees = extract_payroll_from_pdf(args.pdf)
    print(f"{accent('Extracted:')} {important(len(employees))} employees")

    if not employees:
        raise SystemExit("No employee data found in PDF. Check the PDF structure.")

    output_xlsx = args.out or args.pdf.with_suffix(".xlsx")
    print(f"{accent('Creating Excel:')} {output_xlsx}")
    create_excel_with_spacing(employees, output_xlsx)

    print(f"{important('Complete.')} Wrote {len(employees)} employees with totals.")


if __name__ == "__main__":
    main()
